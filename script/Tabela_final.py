import openpyxl
import teste2
import unicodedata
import re


# ---------------------------
# FUNÇÃO PRA NORMALIZAR TEXTO
# (resolve acento, espaço, etc)
# ---------------------------
def normalizar(texto):
    if texto is None:
        return ""

    texto = str(texto).lower().strip()

    # remove acento
    texto = unicodedata.normalize("NFKD", texto).encode("ASCII", "ignore").decode("ASCII")

    # transforma tudo que não é letra ou número em espaço
    texto = re.sub(r'[^a-z0-9]', ' ', texto)

    # remove espaços duplicados
    texto = re.sub(r'\s+', ' ', texto)

    return texto.strip()

# ---------------------------
# ABRIR ARQUIVO
# ---------------------------
arquivo = openpyxl.load_workbook("2026.xlsx")
planilha = arquivo["Dados2026"]

mes = "Fevereiro"
ano = 2026

mes_limpo = normalizar(mes)

# ---------------------------
# DICIONÁRIO (NORMALIZADO)
# ---------------------------
valores = {
    normalizar("Vendas amazon FBA"): teste2.Vendas_amazon_FBA,
    normalizar("Vendas amazon DBA"): teste2.Vendas_amazon_DBA,
    normalizar("Frete pago pelo comprador"): teste2.Frete_pago_pelo_comprador,
    normalizar("credito de embalagem presente"): teste2.credito_de_embalagem_presente,
    normalizar("REEMBOLSO -FBA/DBA- Tarifa de Venda"): teste2.REEMBOLSO_FBADBA_Tarifa_de_Venda,
    normalizar("Reembolso para pacotes extraviados"): teste2.Reembolso_para_pacotes_extraviados,
    normalizar("Tarifa de venda DBA"): teste2.Tarifas_de_venda_DBA,
    normalizar("Tarifa de venda FBA"): teste2.Tarifas_de_venda_FBA,
    normalizar("Tarifa de Manuseio Delivery by amazon"): teste2.Tarifa_de_Manuseio_Delivery_by_amazon,
    normalizar("Estorno do frete pago pelo comprador"): teste2.Estorno_do_frete_pago_pelo_comprador,
    normalizar("PEDIDO-Descontos Promocionais"): teste2.PEDIDO_Descontos_Promocionais,
    normalizar("PEDIDO-IMPOSTO DE VENDAS COLETADAS"): teste2.PEDIDOIMPOSTO_DE_VENDAS_COLETADAS,
    normalizar("REEMBOLSO- FBA/DBA - Credito de remessa"): teste2.REEMBOLSO_FBA_DBA_Credito_de_remessa,
    normalizar("REEMBOLSO- FBA/DBA - Venda do Produto"): teste2.REEMBOLSO_FBADBA_Venda_do_Produto,
    normalizar("Tarifa de devolução de inventário pela Amazon"): teste2.Tarifa_de_devolução_de_inventário_pela_Amazon,
    normalizar("Tarifa de armazenagem do programa FBA"): teste2.Tarifa_de_armazenagem_do_programa_FBA,
    normalizar("Frete da transportadora parceira da Amazon de FBA"): teste2.Frete_transportadora,
    normalizar("Custo de publicidade"): teste2.Custo_publicidade,
    normalizar("Cadastro"): teste2.Cadastro,
    normalizar("Transferencia conta C6"): teste2.Transferencia_conta_C6,
    normalizar("Cobranca no cartao C6"): teste2.Cobranca_no_cartao_C6
}

# ---------------------------
# ENCONTRAR COLUNA DO MÊS
# ---------------------------
coluna_mes = None

for col in range(1, planilha.max_column + 1):

    valor_mes = planilha.cell(row=4, column=col).value

    print(f"Col {col} | Mês: '{valor_mes}'")

    if valor_mes:

        mes_planilha = normalizar(valor_mes)

        if mes_planilha == mes_limpo:
            coluna_mes = col
            print(f"✅ Mês {mes} encontrado na coluna {col}")
            break

if not coluna_mes:
    raise ValueError(f"❌ Mês '{mes}' não encontrado!")

# ---------------------------
# PREENCHER PLANILHA
# ---------------------------
for row in range(6, planilha.max_row + 1):

    descricao_celula = planilha.cell(row=row, column=2).value

    if descricao_celula:

        descricao_normalizada = normalizar(descricao_celula)

        if descricao_normalizada in valores:

            valor_num = float(valores[descricao_normalizada])

            planilha.cell(row=row, column=coluna_mes).value = valor_num
            planilha.cell(row=row, column=coluna_mes).number_format = '#,##0.00'

            print(f"✔ {descricao_celula} → {valor_num}")

# ---------------------------
# SALVAR
# ---------------------------


arquivo.save("2026.xlsx")

print("🚀 Dados inseridos com sucesso!")