import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def limpar_valor_financeiro(v):

    if pd.isna(v):
        return 0.0

    s = str(v).strip()

    # remove separador de milhar
    s = s.replace('.', '')

    # troca vírgula decimal por ponto
    s = s.replace(',', '.')

    try:
        return float(s)
    except:
        return 0.0


def formatar_para_excel(arquivo_csv, arquivo_excel):

    try:

        # 1. Ler CSV (detecção automática de separador)
        df = pd.read_csv(arquivo_csv, sep=None, engine="python")

        # limpar espaços nos nomes das colunas
        df.columns = df.columns.str.strip()

        # ---------------------------
        # MOVER MÊS E ANO PARA O FINAL
        # ---------------------------

        colunas_finais = [c for c in ["Nome do Mês", "Ano"] if c in df.columns]

        outras_colunas = [c for c in df.columns if c not in colunas_finais]

        df = df[outras_colunas + colunas_finais]

        # ---------------------------
        # COLUNAS FINANCEIRAS
        # ---------------------------

        colunas_financeiras = [
            'vendas do produto',
            'créditos de remessa',
            'créditos de embalagem de presente',
            'descontos promocionais',
            'imposto de vendas coletados',
            'tarifas de venda',
            'taxas fba',
            'taxas de outras transações',
            'outro',
            'total'
        ]

        # LIMPAR VALORES
        for col in colunas_financeiras:
            if col in df.columns:
                df[col] = df[col].apply(limpar_valor_financeiro)

        # ---------------------------
        # SALVAR EXCEL
        # ---------------------------

        with pd.ExcelWriter(arquivo_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, index=False, sheet_name='Janeiro')

        wb = load_workbook(arquivo_excel)
        ws = wb['Janeiro']

        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        center_align = Alignment(horizontal='center', vertical='center')

        # ---------------------------
        # FORMATAR CABEÇALHO
        # ---------------------------

        for col_idx, column in enumerate(df.columns, 1):

            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

            larguras = [len(str(valor)) if pd.notna(valor) else 0 for valor in df[column]]

            maior_item = max(larguras) if larguras else 0
            tamanho_titulo = len(str(column))

            largura_final = max(maior_item, tamanho_titulo) + 4

            ws.column_dimensions[get_column_letter(col_idx)].width = largura_final

        # ---------------------------
        # FILTROS
        # ---------------------------

        ultima_coluna = get_column_letter(df.shape[1])
        ws.auto_filter.ref = f"A1:{ultima_coluna}1"

        # congelar cabeçalho
        ws.freeze_panes = "A2"

        # ---------------------------
        # SALVAR
        # ---------------------------

        wb.save(arquivo_excel)

        print(f"✅ Excel gerado com sucesso: {arquivo_excel}")

    except Exception as e:

        print(f"❌ Erro: {e}")


if __name__ == "__main__":

    formatar_para_excel(
        "tabela_final_limpa.csv",
        "2026.xlsx"
    )