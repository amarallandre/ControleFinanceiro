import pandas as pd
import openpyxl

# 1. Configurações de busca (Critérios)
ARQUIVO = '2026.xlsx'
ABA = 'Fevereiro'

criterio_DBA = "Vendedor"
criterio_FBA = "Amazon"
criterio_mes = input("digite o mes ")
criterio_reembolso = "Reembolso"
criterio_pacotes_extraviados = "Reembolso para pacotes extraviados"
criterio_tarifa_de_manuseio = "Tarifa de manuseio com base no peso Delivery by Amazon"
criterio_extorno_do_frete = "Estorno do frete pago pelo comprador"
criterio_tarifa_de_devolucao = "Tarifa de devolução de inventário pela Amazon"
criterio_tarifa_de_armazenagem = "Tarifa de armazenagem do programa FBA - Logística da Amazon"
criterio_frete_da_transportadora = "Frete da transportadora parceira da Amazon de FBA"
criterio_custo_de_publi = "Custo de Publicidade"
criterio_cadastro = "Cadastro"
criterio_transferir = "Transferir"
criterio_debito = "Débito"
criterio_ano = int(input("digite o ano ")) # Se no Excel for número, use sem aspas. Se for texto, use "2026"

# 2. Carregando os dados
try:
    df_dados = pd.read_excel(ARQUIVO, sheet_name=ABA)

    # Limpeza preventiva: remove espaços em branco no início/fim das células de texto
    # Isso evita que "Amazon " seja diferente de "Amazon"
    colunas_texto = ['atendimento', 'Nome do Mês', 'Ano']
    for col in colunas_texto:
        if col in df_dados.columns:
            df_dados[col] = df_dados[col].astype(str).str.strip()

    # 3. Tratamento da coluna de valores (N)
    # Se a coluna N vier como texto (ex: "R$ 10,50"), convertemos para número
    if df_dados['vendas do produto'].dtype == 'object':
        df_dados['vendas do produto'] = (
            df_dados['vendas do produto']
            .replace(r'[R\$\s.]', '', regex=True)  # Remove R$, espaços e pontos de milhar
            .replace(',', '.')  # Troca vírgula decimal por ponto
        )
    df_dados['vendas do produto'] = pd.to_numeric(df_dados['vendas do produto'], errors='coerce').fillna(0)

    df_dados["créditos de remessa"] = pd.to_numeric(
        df_dados["créditos de remessa"], errors="coerce"
    ).fillna(0)

    df_dados["créditos de embalagem de presente"] = pd.to_numeric(
        df_dados["créditos de embalagem de presente"], errors="coerce"
    ).fillna(0)

    df_dados["tarifas de venda"] = pd.to_numeric(
        df_dados["tarifas de venda"], errors="coerce"
    ).fillna(0)

    df_dados["outro"] = pd.to_numeric(
        df_dados["outro"], errors="coerce"
    ).fillna(0)

    df_dados["taxas de outras transações"] = pd.to_numeric(
        df_dados["taxas de outras transações"], errors="coerce"
    ).fillna(0)

    df_dados["descontos promocionais"] = pd.to_numeric(
        df_dados["descontos promocionais"], errors="coerce"
    ).fillna(0)

    df_dados["imposto de vendas coletados"] = pd.to_numeric(
        df_dados["imposto de vendas coletados"], errors="coerce"
    ).fillna(0)

    df_dados["créditos de remessa"] = pd.to_numeric(
        df_dados["créditos de remessa"], errors="coerce"
    ).fillna(0)

    # 4. Aplicação do Filtro (Equivalente ao SUMIFS)
    filtro = (
            (df_dados['atendimento'] == criterio_DBA) &
            (df_dados['Nome do Mês'] == criterio_mes) &
            (df_dados['Ano'] == str(criterio_ano))  # Ajuste conforme o tipo de dado na sua planilha
    )

    filtro2 = (
            (df_dados['atendimento'] == criterio_FBA) &
            (df_dados['Nome do Mês'] == criterio_mes) &
            (df_dados['Ano'] == str(criterio_ano))  # Ajuste conforme o tipo de dado na sua planilha
    )

    filtro3 = (
            (df_dados['Nome do Mês'] == criterio_mes) &
            (df_dados['Ano'] == str(criterio_ano))
    )

    filtro4 = (
            (df_dados['Nome do Mês'] == criterio_mes) &
            (df_dados['Ano'] == str(criterio_ano)) &
            (df_dados['tipo'] == criterio_reembolso)
    )

    filtro5 = (
            (df_dados['Nome do Mês'] == criterio_mes) &
            (df_dados['Ano'] == str(criterio_ano)) &
            (df_dados['descrição'] == criterio_pacotes_extraviados)
    )
    filtro6 = (
            (df_dados['Nome do Mês'] == criterio_mes) &
            (df_dados['Ano'] == str(criterio_ano)) &
            (df_dados['descrição'] == criterio_tarifa_de_manuseio)
    )
    filtro7 = (
            (df_dados['Nome do Mês'] == criterio_mes) &
            (df_dados['Ano'] == str(criterio_ano)) &
            (df_dados['descrição'] == criterio_extorno_do_frete)
    )

    filtro8 = (
            (df_dados['Nome do Mês'] == criterio_mes) &
            (df_dados['Ano'] == str(criterio_ano)) &
            (df_dados['descrição'] == criterio_tarifa_de_devolucao)
    )

    filtro9 = (
            (df_dados['Nome do Mês'] == criterio_mes) &
            (df_dados['Ano'] == str(criterio_ano)) &
            (df_dados['descrição'] == criterio_tarifa_de_armazenagem)
    )

    filtro10 = (
            (df_dados['Nome do Mês'] == criterio_mes) &
            (df_dados['Ano'] == str(criterio_ano)) &
            (df_dados['descrição'] == criterio_frete_da_transportadora)
    )

    filtro11 = (
            (df_dados['Nome do Mês'] == criterio_mes) &
            (df_dados['Ano'] == str(criterio_ano)) &
            (df_dados['descrição'] == criterio_custo_de_publi)
    )

    filtro12 = (
            (df_dados['Nome do Mês'] == criterio_mes) &
            (df_dados['Ano'] == str(criterio_ano)) &
            (df_dados['descrição'] == criterio_cadastro)
    )

    filtro13 = (
            (df_dados['Nome do Mês'] == criterio_mes) &
            (df_dados['Ano'] == str(criterio_ano)) &
            (df_dados['tipo'] == criterio_transferir)
    )

    filtro14 = (
            (df_dados['Nome do Mês'] == criterio_mes) &
            (df_dados['Ano'] == str(criterio_ano)) &
            (df_dados['tipo'] == criterio_debito)
    )




    # 5. Cálculo do Resultado
    Vendas_amazon_FBA = df_dados.loc[filtro, 'vendas do produto'].sum()
    Vendas_amazon_DBA = df_dados.loc[filtro2, 'vendas do produto'].sum()
    Frete_pago_pelo_comprador = df_dados.loc[filtro3, "créditos de remessa"].sum()
    credito_de_embalagem_presente = df_dados.loc[filtro3, "créditos de embalagem de presente"].sum()
    REEMBOLSO_FBADBA_Tarifa_de_Venda = df_dados.loc[filtro4, "tarifas de venda"].sum()
    Reembolso_para_pacotes_extraviados = df_dados.loc[filtro5, "outro"].sum()
    Tarifas_de_venda_DBA = df_dados.loc[filtro, "tarifas de venda"].sum()
    Tarifas_de_venda_FBA = df_dados.loc[filtro2, "tarifas de venda"].sum()
    Tarifa_de_Manuseio_Delivery_by_amazon = df_dados.loc[filtro6, "taxas de outras transações"].sum()
    Estorno_do_frete_pago_pelo_comprador = df_dados.loc[filtro7, "taxas de outras transações"].sum()
    PEDIDO_Descontos_Promocionais = df_dados.loc[filtro3, "descontos promocionais"].sum()
    PEDIDOIMPOSTO_DE_VENDAS_COLETADAS = df_dados["imposto de vendas coletados"].sum()
    REEMBOLSO_FBA_DBA_Credito_de_remessa = df_dados.loc[filtro4, "créditos de remessa"].sum()
    REEMBOLSO_FBADBA_Venda_do_Produto = df_dados.loc[filtro4, "vendas do produto"].sum()
    Tarifa_de_devolução_de_inventário_pela_Amazon = df_dados.loc[filtro8, "outro"].sum()
    Tarifa_de_armazenagem_do_programa_FBA = df_dados.loc[filtro9, "outro"].sum()
    Frete_transportadora = df_dados.loc[filtro10, "outro"].sum()
    Custo_publicidade = df_dados.loc[filtro11, "taxas de outras transações"].sum()
    Cadastro = df_dados.loc[filtro12, "outro"].sum()
    Transferencia_conta_C6 = df_dados.loc[filtro13, "outro"].sum()
    Cobranca_no_cartao_C6 = df_dados.loc[filtro14, "outro"].sum()

    TOTAL_RECEBIDO_AMAZON = Vendas_amazon_FBA + Vendas_amazon_DBA + Frete_pago_pelo_comprador + credito_de_embalagem_presente + REEMBOLSO_FBADBA_Tarifa_de_Venda + Reembolso_para_pacotes_extraviados
    TOTAL_DESCONTO_AMAZON = Tarifas_de_venda_DBA + Tarifas_de_venda_FBA + Tarifa_de_Manuseio_Delivery_by_amazon + Estorno_do_frete_pago_pelo_comprador + PEDIDO_Descontos_Promocionais + PEDIDOIMPOSTO_DE_VENDAS_COLETADAS + REEMBOLSO_FBA_DBA_Credito_de_remessa  + REEMBOLSO_FBADBA_Venda_do_Produto  + Tarifa_de_devolução_de_inventário_pela_Amazon  + Tarifa_de_armazenagem_do_programa_FBA + Frete_transportadora + Custo_publicidade
    TOTAL_JÁ_INCLUIDO_OS_DESCONTOS = TOTAL_RECEBIDO_AMAZON + TOTAL_DESCONTO_AMAZON

    valores = {
        "Vendas amazon FBA": Vendas_amazon_FBA,
        "Vendas amazon DBA": Vendas_amazon_DBA,
        "Frete pago pelo  comprador": Frete_pago_pelo_comprador,
        "credito de embalagem presente": credito_de_embalagem_presente,
        "REEMBOLSO -FBA/DBA- Tarifa de Venda": REEMBOLSO_FBADBA_Tarifa_de_Venda,
        "Reembolso para pacotes extraviados": Reembolso_para_pacotes_extraviados,
        "TOTAL RECEBIDO AMAZON": TOTAL_RECEBIDO_AMAZON,
        "Tarifas de venda DBA": Tarifas_de_venda_DBA,
        "Tarifas de venda FBA": Tarifas_de_venda_FBA,
        "Tarifa de Manuseio Delivery by amazon": Tarifa_de_Manuseio_Delivery_by_amazon,
        "Estorno do frete pago pelo comprador": Estorno_do_frete_pago_pelo_comprador,
        "PEDIDO-Descontos Promocionais": PEDIDO_Descontos_Promocionais,
        "PEDIDO-IMPOSTO DE VENDAS COLETADAS": PEDIDOIMPOSTO_DE_VENDAS_COLETADAS,
        "REEMBOLSO- FBA/DBA - Credito de remessa": REEMBOLSO_FBA_DBA_Credito_de_remessa,
        "REEMBOLSO- FBA/DBA - Venda do Produto": REEMBOLSO_FBADBA_Venda_do_Produto,
        "Tarifa de devolução de inventário pela Amazon": Tarifa_de_devolução_de_inventário_pela_Amazon,
        "Tarifa de armazenagem do programa FBA": Tarifa_de_armazenagem_do_programa_FBA,
        "Frete transportadora": Frete_transportadora,
        "Custo publicidade": Custo_publicidade,
        "Cadastro": Cadastro,
        "TOTAL DESCONTO AMAZON": TOTAL_DESCONTO_AMAZON,
        "TOTAL JÁ INCLUIDO OS DESCONTOS": TOTAL_JÁ_INCLUIDO_OS_DESCONTOS,
        "Transferencia conta C6": Transferencia_conta_C6,
        "Cobranca no cartao C6": Cobranca_no_cartao_C6
    }

    arquivo = openpyxl.load_workbook(ARQUIVO)
    planilha = arquivo["Dados2026"]

    try:
        # 1. Padroniza o mês para a busca
        mes_procurado = criterio_mes.strip().lower()
        # 2. Define o ano que você está buscando (ajuste conforme sua variável de ano)
        ano_procurado = str(criterio_ano).strip()

        # 3. Validação do Ano (Lendo da célula A3, conforme sua imagem)
        ano_na_planilha = str(planilha["A3"].value).strip()

        if ano_na_planilha != ano_procurado:
            print(f"Atenção: A planilha está marcada como {ano_na_planilha}, mas você quer o ano {ano_procurado}.")
            # Aqui você decide se para o código ou continua.
            # Vou colocar um sys.exit() ou return se estiver em função, ou apenas um aviso.

        else:
            # Encontrar coluna do mês (Linha 4 conforme a imagem)
            coluna_mes = None
            for col in range(1, planilha.max_column + 1):
                valor = planilha.cell(row=4, column=col).value
                if valor and str(valor).strip().lower() == mes_procurado:
                    coluna_mes = col
                    break

            if coluna_mes is None:
                print(f"Mês '{criterio_mes}' não encontrado na linha 4 para o ano {ano_procurado}.")
            else:
                # Preencher os valores
                for row in range(6, planilha.max_row + 1):
                    # Pegando a descrição da coluna 1 (A)
                    descricao_celula = planilha.cell(row=row, column=3).value

                    if descricao_celula:
                        descricao = str(descricao_celula).strip()

                        if descricao in valores:
                            celula = planilha.cell(row=row, column=coluna_mes)

                            # Converte para float e trata possíveis erros de valor
                            try:
                                valor_num = float(valores[descricao])
                                celula.value = valor_num
                                celula.number_format = '#,##0.00'
                            except (ValueError, TypeError):
                                print(f"Erro ao converter valor para {descricao}")

                arquivo.save(ARQUIVO)
                print(f"Tabela de {criterio_mes}/{ano_procurado} atualizada com sucesso!")



    except FileNotFoundError:
        print(f"Erro: O arquivo '{ARQUIVO}' não foi encontrado.")
except KeyError as e:
    print(f"Erro: A coluna {e} não existe na aba '{ABA}'. Verifique os cabeçalhos.")
except Exception as e:
    print(f"Ocorreu um erro inesperado: {e}")